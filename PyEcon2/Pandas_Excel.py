# -*- coding: utf-8 -*-

import pandas as pd
import openpyxl as opx


book = r"GDP.xlsx"    # xlsxファイルの保存場所
sheet= "sheet1"    # xlsxの中から読み込みたいシート名を選択
EXL  = pd.ExcelFile(book)    # xlsxファイルをPython上で開く
Data = EXL.parse(sheet,index_col="TIME") # 開いたxlsxファイルからデータの入ったシートを時系列のDataFrame形式で読み込む




Data.describe()    # 記述統計量を出力
Data.plot()    # チャートを出力



desc = Data.describe()  # 記述統計量一覧のDataFrameオブジェクトを作る
outpath =  r"out.xlsx"  #　保存先のファイルパス
desc.to_excel(outpath,"sheet1") # 新規xlsxファイルのsheet1に保存するメソッド



outpath =  r"GDP.xlsx" # 保存先のファイルパス＝入力に使用したGDPのファイル
newsheetname = "out" # 追加するシート名






"""
既存のxlsxファイルに新しいシートを追加する形でDataFrameを保存するには、openpyxlというパッケージのload_workbookという関数を使う必要があります。こちらもAnacondaに含まれていますので、importすることですぐに使えます。

試しに、記述統計の一覧表descを、既存のGDP．xlsxにOutという新規シートを追加して保存してみます。
"""



desc = pd.DataFrame(random(100)).describe()


outpath= r"Book.xlsx"



newsheetname="Out" # 新規シート名

# 既存のシートを読み込み、書き込みリストに登録
book = opx.load_workbook(outpath)
writer = pd.ExcelWriter(outpath)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

# 分析結果を新規シートに書き込み登録
desc.to_excel(writer, newsheetname)

# 書き込み
writer.save()




"""
最後に、既存のシートの特定のセルに計算した結果を入力する方法を書いておきます。例えば、GDP.xlsx、tableシートにあらかじめ作ってあるテーブルがあり、そのセルB2~B9に記述統計の数値を入力する作業を行ってみます。
"""

editpath= r"Book.xlsx"
editsheetname="table" # 既存シート名


# 既存のシートを読み込み、対象セルの値を上書き
book = opx.load_workbook(editpath)
ws = book.worksheets[book.get_sheet_names().index(editsheetname)]


# 2~9の連番で繰り返し。DataFrameのから入力したい数値を呼び出し、対応するセルの値を上書き
for i in range(2,10):
    ws.cell("B"+str(i)).value = desc.iloc[i-2,0]

# 変更をGDP.ｘｌｓｘに上書き保存
book.save(editpath)


"""
以上の書き出し方を組み合わせれば、比較的自由にPythonとExcelの連携ができると思います。

例で扱った操作は、わざわざPythonを使う必要のない単純なものでしたが、ある程度大きなデータを繰り返し操作するようなルーティンワークや、本格的なデータ分析の結果をExcelで提出するには役立つと思います。

今後も紹介していきますように、Pythonをベースにして使える機能は、ExcelのVBAを書くよりも圧倒的に豊富ですので、普段Excelを多用している実務家の方にもぜひお勧めしたいと思います。

"""




